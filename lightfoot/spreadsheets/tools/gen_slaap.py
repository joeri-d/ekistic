"""Lightfoot — slaapsysteem.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/slaapsysteem.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "SLAAP — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# SLAAP",
        "",
        "## Ontwerpdoel",
        "Een slaapopstelling die diepe, herstellende slaap garandeert met minimale investering,",
        "minimale onderhoudslast en maximale levensduur van de componenten.",
        "",
        "## Standaardnorm",
        "→ Basis: metalen lattenbodem of latten eenvoudig houten frame (IKEA Neiden 90×200).",
        "→ Matras: medium-firm schuim/hybride — Emma Original of gelijkwaardig.",
        "→ Matrasbeschermer: waterdicht, ademend (Protect-A-Bed of wit-label equivalent).",
        "→ Textiel: 2 sets percale katoen (200TC+), 1 all-season dekbed, 2 dekbedovertrekken, 2 kussenslopen.",
        "→ Kussen: 1 ergonomisch middenhard kussen (zij- of rugslaper norm).",
        "→ Afmeting: 90×200 cm als standaard voor single-persoon.",
        "",
        "## Uitzondering (niet standaard)",
        "→ Hellende onderlegger / wedge: uitsluitend bij gedocumenteerde reflux of positionele apneu.",
        "→ Verstelbare latenbodem: niet nodig — statische flat base voldoet voor de overgrote meerderheid.",
        "→ Elektrische deken: niet opgenomen — betere ventilatiegewoonte primeert.",
        "",
        "## Onderhoud",
        "→ Matrasbeschermer: wassen 60° elke 2 maanden.",
        "→ Lakens + slopen: wassen 60° wekelijks.",
        "→ Dekbedovertrek: wassen 40° om de 2 weken.",
        "→ Matras: keren (180°) elke 6 maanden — verlengt levensduur.",
        "→ Frame: jaarlijkse check schroeven en latten.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [30, 34, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Slaapsysteem 90×200")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        ("Bedframe / lattenbodem", "Metaal, 90×200 cm", "IKEA Neiden", 1, 59, 120),
        ("Matras", "Medium-firm, koudschuim + pocket, 90×200", "Emma Original", 1, 379, 120),
        ("Matrasbeschermer", "Waterdicht, 90×200, ademend", "Protect-A-Bed Basic", 1, 45, 60),
        ("Laken (hoeslaken)", "Percale katoen 200TC, wit", "IKEA Dvala", 2, 18, 48),
        ("Dekbed all-season", "200 g/m² + 100 g/m² combinatie, 140×200", "IKEA Fjällbräcka", 1, 55, 84),
        ("Dekbedovertrek", "Percale katoen, wit, 140×200", "IKEA Dvala", 2, 22, 60),
        ("Kussen", "Middenhard, ergonomisch, 60×70", "Emma Comfort Pillow", 1, 49, 36),
        ("Kussenslopen", "Percale katoen, wit, 60×70", "IKEA Dvala", 2, 8, 48),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def verbruik(wb):
    ws = mk_ws(wb, "Verbruiksgoederen", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 16])
    title_block(ws, 1, "Verbruiksgoederen — Slaap")
    hdr(ws, 3, ["Item", "Eenheid", "Qty / jaar", "Prijs/eenheid (€)",
                "Jaarkost (€)", "Maandkost (€)"])

    items = [
        ("Wasmiddel textiel (60°)", "fles 1.5L", 3, 8),
        ("Wasverzachter", "fles 1L", 2, 5),
        ("Anti-allergeen spray matras", "bus 400ml", 1, 9),
    ]

    r = 4
    totaal_jaar = 0
    for item, eenheid, qty, prijs in items:
        jaar = qty * prijs
        maand = round(jaar / 12, 2)
        totaal_jaar += jaar
        drow(ws, r, [item, eenheid, qty, prijs, jaar, maand],
             fmts=[None, None, None, EUR, EUR0, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", totaal_jaar, round(totaal_jaar / 12, 2)])
    ws.cell(r, 5).number_format = EUR0
    ws.cell(r, 6).number_format = EUR

def onderhoud(wb):
    ws = mk_ws(wb, "Onderhoud", "1A2B4A")
    set_cols(ws, [28, 34, 16, 12, 24, 14])
    title_block(ws, 1, "Onderhoudsblad — Slaap")
    hdr(ws, 3, ["Item", "Taak", "Frequentie", "Duur (min)", "Materialen", "Jaarkosten (€)"])

    taken = [
        ("Hoeslakens + kussenslopen", "Was 60°, ophangen of droger", "Wekelijks", 15, "Wasmiddel", 20),
        ("Dekbedovertrek", "Was 40°, ophangen", "2× / maand", 15, "Wasmiddel", 8),
        ("Matrasbeschermer", "Was 60°", "1× / 2 maanden", 10, "Wasmiddel", 6),
        ("Matras", "Keren 180°", "2× / jaar", 10, "-", 0),
        ("Matras", "Luchtend ophangen", "1× / jaar", 20, "-", 0),
        ("Bedframe", "Check schroeven, latten reinigen", "1× / jaar", 15, "-", 0),
        ("Kussen", "Aireo + vlekcheck", "1× / 3 maanden", 5, "Anti-allergeen spray", 9),
    ]

    r = 4
    for groep, taak, freq, duur, mat, jk in taken:
        drow(ws, r, [groep, taak, freq, duur, mat, jk],
             fmts=[None, None, None, None, None, EUR0])
        r += 1

    trow(ws, r, ["TOTAAL JAARKOSTEN", "", "", "", "", sum(t[5] for t in taken)])
    ws.cell(r, 6).number_format = EUR0

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Slaap (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Volledig slaapsysteem (8 items)", 699, "", "Eenmalige aankoopkost"])
    trow(ws, 6, ["Totaal investering", 699, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Afschrijving slaapsysteem", 8.35, "Op basis van levensduur per item"),
        ("Verbruiksgoederen (was)", 2.92, "Wasmiddel, verzachter, spray"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+3, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Lakens wassen + ophangen", 60, ""),
        ("Dekbedovertrek wassen", 30, ""),
        ("Matrasbeschermer wassen", 10, ""),
        ("Matras keren / luchten", 5, ""),
    ]
    r2 = r + 4
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Slaap")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Slaap", "Wonen / Shelter", "Ruimte voor 90×200 bed + 60 cm looppad rondom vereist."),
        ("Slaap", "Hygiëne", "Was textiel 60° — wasmachine of wasplaats in woning vereist."),
        ("Slaap", "Wonen / Shelter", "Verwarmde slaapruimte: 16–18 °C norm voor slaapkwaliteit."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    verbruik(wb)
    onderhoud(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
