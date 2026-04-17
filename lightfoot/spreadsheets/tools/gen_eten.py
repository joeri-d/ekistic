"""Lightfoot — eetsysteem.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/eetsysteem.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "ETEN — Filosofie & Voedingsprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# EETSYSTEEM",
        "",
        "## Ontwerpdoel",
        "Een vast, sober, herhaalbaar voedingssysteem dat gezond is (laag LDL, volledige macro's",
        "en micro's), goedkoop, weinig tijdintensief en vrij van culinaire keuzestress.",
        "",
        "## Voedingsprincipes",
        "→ Laag LDL: geen verzadigd vet uit vlees of zuivel als basis. Vis 1× per week is uitzondering.",
        "→ Eiwitten: peulvruchten (linzen, kikkererwten, bonen) + ei (3-4/week) + tempeh.",
        "→ Koolhydraten: volkoren (haver, rijst, quinoa, volkoren brood) — lage GI.",
        "→ Vetten: olijfolie, noten, avocado (1× per week) — onverzadigd.",
        "→ Vezels: 30g/dag via groenten, peulvruchten, haver.",
        "→ Micronutriënten: B12-supplement (vereist bij plantaardig) + D3 in winter.",
        "→ Water is de normdrank — 1.5-2L/dag.",
        "",
        "## Maaltijdstructuur (vast schema)",
        "→ Ontbijt: overnight oats + bevroren bessen + banaan + gemengde noten.",
        "→ Lunch: volkoren boterham (2st) + hummus of notenpasta + rauwkost.",
        "→ Diner: zie 5-daags rotatiesysteem hieronder.",
        "→ Tussendoortje: fruit (1 stuk) of noten (30g) — optioneel.",
        "",
        "## Dinerrotatie (5-daags, herhaalbaar)",
        "→ Dag 1: rode linzendal + bruine rijst + spinazie.",
        "→ Dag 2: zwarte bonen + zoete aardappel + broccoli.",
        "→ Dag 3: kikkererwten stoofpot (chermoula) + couscous.",
        "→ Dag 4: tempeh roerbak + volkoren noedels + paprika/courgette.",
        "→ Dag 5: omelet (2 eieren) + groenten + volkoren brood.",
        "",
        "## Inkoopstrategie",
        "→ 1× per week inkopen (vast dag, vaste lijst — geen impulsaankopen).",
        "→ Seizoensgroenten voor kostenoptimalisatie.",
        "→ Droogwaren (linzen, rijst, havermout) in bulk — 1× per maand.",
        "→ Diepvries: bessen, groenten — kwaliteit op prijs gebufferd.",
        "",
        "## Uitzondering",
        "→ Vis (wilde zalm, makreel): 1× per week ter optimalisatie omega-3.",
        "→ Restaurants / eetgelegenheid buiten: budgettair niet opgenomen in model.",
    ])

def inventaris_keuken(wb):
    ws = mk_ws(wb, "Inventaris keuken", "2D4A7A")
    set_cols(ws, [28, 32, 20, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Keukenmateriaal")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        ("Inductiekookplaat", "2-pits, 3500W, compact", "Tefal IH720812", 1, 79, 120),
        ("Pot 3L", "Roestvrij staal, deksel inbegrepen", "IKEA 365+", 1, 25, 180),
        ("Braadpan 28cm", "Keramisch anti-aanbak, geen PTFE", "GreenPan Venice", 1, 45, 48),
        ("Deksel universeel", "Glas, 26-28cm", "IKEA 365+", 1, 8, 120),
        ("Koksmes 20cm", "Roestvrij, gesmeed", "Victorinox Fibrox", 1, 35, 240),
        ("Snijplank", "PP-kunststof, vaatwasmachinebestendig", "IKEA Legitim", 1, 8, 60),
        ("Vergiet / zeef", "RVS, 22cm", "IKEA Idealisk", 1, 6, 120),
        ("Houten lepel + spatel set", "Bamboe", "IKEA Rotera", 1, 5, 36),
        ("Maatbeker 1L", "Glas, markering ml/cups", "IKEA Vardagen", 1, 4, 120),
        ("Blikopener", "Manueel, RVS", "Zyliss Smooth Edge", 1, 12, 120),
        ("Diep bord", "Porselein, wit", "IKEA Vardagen", 2, 4, 120),
        ("Plat bord", "Porselein, wit", "IKEA Vardagen", 2, 3, 120),
        ("Soepkom / kom", "Porselein, wit", "IKEA Vardagen", 1, 3, 120),
        ("Bestek (vork, mes, lepel, theelepel)", "RVS, eenvoudig", "IKEA Mopsig", 1, 10, 120),
        ("Glas (water)", "Gehard glas, 30cl", "IKEA Pokal", 2, 2, 60),
        ("Mengkom", "RVS, 3L", "IKEA Blanda Blank", 1, 7, 120),
        ("Keukenweegschaal", "Digitaal, 5kg/1g", "Beurer KS19", 1, 15, 120),
        ("Waterkoker", "1.7L, RVS", "Russell Hobbs 21370", 1, 35, 84),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def voeding_budget(wb):
    ws = mk_ws(wb, "Voedingsbudget", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 18])
    title_block(ws, 1, "Voedingsbudget — Maandelijks (1 persoon)")
    hdr(ws, 3, ["Categorie", "Qty / maand", "Eenheid", "Prijs/eenheid (€)",
                "Maandkost (€)", "Toelichting"])

    voeding = [
        # (categorie, qty, eenheid, prijs, toelichting)
        ("Havermout (droog)", 1.5, "kg", 1.80, "Ontbijt basis"),
        ("Bevroren bessen mix", 2, "kg", 4.50, "Ontbijt"),
        ("Bananen", 8, "stuk", 0.25, "Ontbijt / tussendoortje"),
        ("Gemengde noten (ongezouten)", 0.5, "kg", 9.00, "Ontbijt + snack"),
        ("Volkoren brood", 4, "brood", 2.50, "Lunch"),
        ("Hummus (pot 400g)", 4, "pot", 2.20, "Lunch"),
        ("Rauwkost (wortelen, komkommer, paprika)", 2, "kg", 2.00, "Lunch"),
        ("Rode linzen (droog)", 1, "kg", 2.50, "Diner D1"),
        ("Zwarte bonen (blik)", 4, "blik", 0.90, "Diner D2"),
        ("Zoete aardappelen", 1.5, "kg", 2.00, "Diner D2"),
        ("Kikkererwten (blik)", 4, "blik", 0.90, "Diner D3"),
        ("Couscous (droog)", 0.5, "kg", 2.00, "Diner D3"),
        ("Tempeh", 4, "blok 200g", 2.50, "Diner D4"),
        ("Volkoren noedels", 0.5, "kg", 2.00, "Diner D4"),
        ("Eieren", 12, "stuk", 0.28, "Diner D5 + ontbijt"),
        ("Bruine rijst (droog)", 1, "kg", 2.20, "Diner D1"),
        ("Gemengde groenten (vers)", 3, "kg", 2.50, "Diner basis"),
        ("Spinazie (diepvries)", 1, "kg", 3.00, "Diner D1"),
        ("Broccoli (vers/diepvries)", 1, "kg", 2.00, "Diner D2"),
        ("Olijfolie extra vergine", 0.5, "L", 8.00, "Kookvet"),
        ("Tomatenpuree (blik)", 4, "blik 400g", 0.80, "Basis sausen"),
        ("Kruiden en specerijen (mix)", 1, "maand", 5.00, "Vaste basis"),
        ("Wilde zalm / makreel (blik)", 4, "blik", 2.20, "Omega-3 wekelijks"),
        ("D3 + B12 supplement", 1, "maand", 8.00, "Micronutriënten"),
        ("Koffie (bonen)", 0.25, "kg", 14.00, "Optioneel — niet meegerekend als norm"),
    ]

    r = 4
    totaal = 0
    for cat, qty, eenheid, prijs, toel in voeding:
        maandkost = round(qty * prijs, 2)
        totaal += maandkost
        bg = C["amber"] if "Optioneel" in toel else "alt"
        drow(ws, r, [cat, qty, eenheid, prijs, maandkost, toel],
             bg=bg,
             fmts=[None, None, None, EUR, EUR, None])
        r += 1

    trow(ws, r, ["TOTAAL VOEDING / MAAND (excl. koffie)", "", "", "", round(totaal - 3.50, 2), ""])
    ws.cell(r, 5).number_format = EUR

def onderhoud(wb):
    ws = mk_ws(wb, "Onderhoud", "1A2B4A")
    set_cols(ws, [28, 34, 16, 12, 24, 14])
    title_block(ws, 1, "Onderhoudsblad — Keuken & Eetsysteem")
    hdr(ws, 3, ["Item / Groep", "Taak", "Frequentie", "Duur (min)", "Materialen", "Jaarkosten (€)"])

    taken = [
        ("Borden, bestek, glazen", "Afwas met warm water + afwasmiddel", "Dagelijks", 10, "Afwasmiddel, spons", 20),
        ("Pan (keramisch)", "Koud laten afkoelen, afwassen met zacht spons", "Na elk gebruik", 5, "Zacht afwasmiddel", 5),
        ("Pot RVS", "Afwassen, droogwrijven", "Na elk gebruik", 5, "Afwasmiddel", 0),
        ("Snijplank", "Afwassen 60° (vaatwasser)", "2× / week", 3, "Afwasmiddel", 0),
        ("Keukenweegschaal", "Afnemen met vochtige doek", "Wekelijks", 2, "-", 0),
        ("Inductieplaat", "Afnemen met vochtige doek", "Wekelijks", 5, "Keukenreiniger", 8),
        ("Waterkoker", "Ontkalken", "1× / 3 maanden", 10, "Citroenzuur", 4),
        ("Snijplank vervanging", "Vervangen bij inkervingen", "1× / 5 jaar", "-", "Nieuwe plank", 2),
        ("Keuken algemeen", "Grondig reinigen kast + werkblad", "Maandelijks", 20, "Ontvetter", 12),
    ]

    r = 4
    for groep, taak, freq, duur, mat, jk in taken:
        drow(ws, r, [groep, taak, freq, duur, mat, jk],
             fmts=[None, None, None, None, None, EUR0])
        r += 1

    trow(ws, r, ["TOTAAL JAARKOSTEN KEUKENONDERHOUD", "", "", "", "", sum(t[5] for t in taken)])
    ws.cell(r, 6).number_format = EUR0

def verbruik(wb):
    ws = mk_ws(wb, "Verbruiksgoederen", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 16])
    title_block(ws, 1, "Verbruiksgoederen — Keuken (niet-voeding)")
    hdr(ws, 3, ["Item", "Eenheid", "Qty / jaar", "Prijs/eenheid (€)",
                "Jaarkost (€)", "Maandkost (€)"])

    items = [
        ("Afwasmiddel (geconcentreerd)", "fles 500ml", 6, 2.50),
        ("Afwasborstels / spons", "stuk", 12, 0.80),
        ("Keukenreiniger spray", "bus 750ml", 3, 3.50),
        ("Citroenzuur (ontkalker)", "zakje 100g", 4, 1.20),
        ("Keukenpapier", "rol", 12, 1.00),
        ("Keukenhanddoek (theedoek)", "stuk — jaarlijks vervangen", 3, 4.00),
        ("Alufolie / bakpapier", "rol", 2, 3.00),
    ]

    r = 4
    totaal_jaar = 0
    for item, eenheid, qty, prijs in items:
        jaar = round(qty * prijs, 2)
        maand = round(jaar / 12, 2)
        totaal_jaar += jaar
        drow(ws, r, [item, eenheid, qty, prijs, jaar, maand],
             fmts=[None, None, None, EUR, EUR0, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", round(totaal_jaar, 2), round(totaal_jaar / 12, 2)])
    ws.cell(r, 5).number_format = EUR0
    ws.cell(r, 6).number_format = EUR

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Eten (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING KEUKEN").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Keukenmateriaal volledig (18 items)", 306, "", "Eenmalige aankoopkost"])
    trow(ws, 6, ["Totaal investering keuken", 306, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Voeding (norm, excl. koffie)", 163.00, "Zie tabblad Voedingsbudget"),
        ("Keukenonderhoud verbruiksgoederen", 3.58, "Afwasmiddel, reiniger, etc."),
        ("Afschrijving keukenmateriaal", 3.90, "Op basis van levensduur per item"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+3, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Weekinkopen + opbergen", 90, ""),
        ("Maandinkoop droogwaren (bulk)", 30, ""),
        ("Koken (dagelijks, gemiddeld)", 300, ""),
        ("Afwas / keuken opruimen", 200, ""),
        ("Maaltijdprep (batchen)", 60, ""),
    ]
    r2 = r + 4
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Eten")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Eten", "Wonen / Shelter", "Keuken met koelkast (min. 60L), werkblad, stopcontact voor inductie vereist."),
        ("Eten", "Wonen / Shelter", "Diepvriesvak min. 30L voor bessen en groenten in bulk."),
        ("Eten", "Hygiëne", "Afwas deelt water-infrastructuur met hygiënefunctie."),
        ("Eten", "Werk", "Lunchpauzeplanning afgestemd op werkritme."),
        ("Hygiëne", "Eten", "Handen wassen vóór koken deelt wastafel/water met hygiëne."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris_keuken(wb)
    voeding_budget(wb)
    onderhoud(wb)
    verbruik(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
