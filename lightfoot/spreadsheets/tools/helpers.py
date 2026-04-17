"""Lightfoot — gedeelde stijlhulpfuncties voor Excel-generatie."""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    "navy":     "1A2B4A",
    "navy_mid": "2D4A7A",
    "grey":     "F5F7FA",
    "amber":    "FFF8E7",
    "blue":     "E8F0FE",
    "green":    "E8F5E9",
    "white":    "FFFFFF",
    "black":    "1A1A1A",
}

_th = Side(style="thin", color="CCCCCC")

def F(c): return PatternFill("solid", fgColor=c)
def fH(s=10): return Font(name="Calibri", bold=True, color="FFFFFF", size=s)
def fB(s=10, col="1A1A1A"): return Font(name="Calibri", bold=True, color=col, size=s)
def fS(s=10): return Font(name="Calibri", size=s, color="1A1A1A")
def fI(s=9): return Font(name="Calibri", italic=True, size=s, color="666666")
def aC(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def aL(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def aR(): return Alignment(horizontal="right", vertical="center")
def Bo(): return Border(left=_th, right=_th, top=_th, bottom=_th)

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr(ws, r, vals, bg=None):
    bg = bg or C["navy"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.fill = F(bg); cell.font = fH(); cell.alignment = aC(); cell.border = Bo()

def sec(ws, r, text, ncols, bg=None):
    bg = bg or C["navy_mid"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(r, 1, text)
    cell.fill = F(bg); cell.font = fB(col="FFFFFF"); cell.alignment = aL(); cell.border = Bo()

def drow(ws, r, vals, bg="alt", fmts=None):
    row_bg = C["grey"] if r % 2 == 0 else C["white"]
    if bg != "alt":
        row_bg = bg
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.fill = F(row_bg)
        cell.font = fS()
        cell.alignment = aL() if c == 1 else (aR() if isinstance(v, (int, float)) else aC())
        cell.border = Bo()
        if fmts and fmts[c-1]:
            cell.number_format = fmts[c-1]

def trow(ws, r, vals, bg=None):
    bg = bg or C["blue"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.fill = F(bg); cell.font = fB(); cell.alignment = aC(); cell.border = Bo()

def mk_ws(wb, name, tab_color=None):
    ws = wb.create_sheet(name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    return ws

def title_block(ws, r, t, sub=""):
    ws.row_dimensions[r].height = 28
    cell = ws.cell(r, 1, t)
    cell.font = Font(name="Calibri", bold=True, size=14, color=C["navy"])
    if sub:
        ws.row_dimensions[r+1].height = 16
        ws.cell(r+1, 1, sub).font = fI()

def nota_blok(ws, start_row, lijnen):
    """Schrijf een blok vrije tekst (notities/filosofie) in kolom A."""
    for i, lijn in enumerate(lijnen):
        r = start_row + i
        ws.row_dimensions[r].height = 18 if lijn else 6
        cell = ws.cell(r, 1, lijn)
        if lijn.startswith("##"):
            cell.font = fB(s=11, col=C["navy"])
        elif lijn.startswith("#"):
            cell.font = fB(s=13, col=C["navy"])
        elif lijn.startswith("→") or lijn.startswith("•"):
            cell.font = fS(s=10)
            cell.alignment = aL()
        else:
            cell.font = fS(s=10)
            cell.alignment = aL()
