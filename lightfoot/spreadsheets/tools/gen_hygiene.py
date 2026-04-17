"""Lightfoot — hygienesysteem.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/hygienesysteem.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "HYGIËNE — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# HYGIËNESYSTEEM",
        "",
        "## Ontwerpdoel",
        "Volledige persoonlijke hygiëne met minimaal materiaal, minimale chemische belasting,",
        "lage maandkost en een herhaalbaar dagelijks + wekelijks ritme.",
        "",
        "## Kernprincipes",
        "→ Douche: dagelijks of om de dag, max. 5-7 minuten, lauwwarm water.",
        "→ Zeep: Marseille-blokzeep of equivalent — geen synthetische parfums.",
        "→ Haar: geen shampoo elke dag (co-wash of waterroutine). Kapperbezoek om de 6 weken.",
        "→ Tanden: elektrische tandenborstel + interdentaal borstels + floss — 2× per dag.",
        "→ Scheren: veiligheidsscheerapparaat (DE safety razor) — schoner, goedkoper, minder afval.",
        "→ Huid: SPF50 dagcrème elke ochtend op gelaat — primaire anti-aging en kankerpreventie.",
        "→ Nagels: knippen 1× per 2 weken.",
        "→ Deodorant: kristaldeodorant (aluin) of alcohol-based — geen aluminium-zout antiperspiant.",
        "",
        "## Efficiëntie via afhankelijkheden",
        "→ Tanden poetsen kan onder de douche of aan de afwasbak — geen apart tijdsblok nodig.",
        "→ Scheren na douche (warme huid, open poriën) — betere kwaliteit, minder irritatie.",
        "→ Alle hygiënetaken clusteren in 1 ochtend- en 1 avondroutine.",
        "",
        "## Uitzondering",
        "→ Thuiskapsel (Wahl clipper): uitzondering als alternatief voor kapperbezoek.",
        "   Investering: €35 eenmalig. Breekt even bij na 2 kapperbezoeken.",
        "→ Zonnebrandcrème lichaam (SPF50): alleen in zomer en bij zonblootstelling.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [30, 34, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Hygiëne")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        # (item, spec, merk, qty, prijs, levmnd)
        ("Elektrische tandenborstel", "Sonische technologie, USB laadbaar", "Oclean X Pro", 1, 55, 60),
        ("Opzetborstels (vervanging)", "Compatibel Oclean", "Oclean standaard", 1, 16, 6),
        ("Veiligheidsscheermes (DE)", "Gesloten kam, massief messing", "Merkur 34C", 1, 45, 240),
        ("Scheermeshouder / standaard", "Chroom, wandmontage", "Generiek RVS", 1, 8, 120),
        ("Handdoek groot (lichaam)", "500 g/m², 70×140 cm, wit", "IKEA Njutning", 2, 12, 48),
        ("Handdoek klein (gezicht/handen)", "500 g/m², 40×70 cm", "IKEA Njutning", 2, 6, 36),
        ("Nagelset (knipper + vijl)", "RVS, compact etui", "Zwilling", 1, 18, 120),
        ("Scheerspiegel", "Muurmontage, x5 vergrotend", "Generiek", 1, 15, 120),
        ("Douchegel dispenser", "Wandmontage, navulbaar, 300ml", "Generiek", 1, 12, 60),
        ("Haarborstel / kam", "Anti-statisch", "Tangle Teezer compact", 1, 12, 36),
        ("Wahl thuisknipper (uitzondering)", "Volledig set + opzetkammen", "Wahl Colour Pro", 1, 35, 120),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        bg = C["amber"] if "uitzondering" in item.lower() else "alt"
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             bg=bg,
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def verbruik(wb):
    ws = mk_ws(wb, "Verbruiksgoederen", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 16])
    title_block(ws, 1, "Verbruiksgoederen — Hygiëne (maandelijks)")
    hdr(ws, 3, ["Item", "Eenheid", "Qty / maand", "Prijs/eenheid (€)",
                "Maandkost (€)", "Toelichting"])

    items = [
        ("Marseille blokzeep (lichaam + handen)", "blok 100g", 1.5, 2.00, "Douche + afwas handen"),
        ("Shampoobar of co-wash", "blok 60g", 0.5, 6.00, "Haar, 1× per 2-3 dagen"),
        ("DE scheermesjes (Astra Superior)", "blok 5 stuks", 1, 1.20, "1 mes per week"),
        ("Scheercrème / scheerzeep", "stuk 100g", 0.33, 7.50, "Elke 3 maanden 1 stuk"),
        ("Tandpasta fluor", "tube 75ml", 0.5, 2.50, "1 tube per 2 maanden"),
        ("Interdentaalborstel (ISO 1-2)", "pak 8st", 1, 2.80, "Dagelijks gebruik"),
        ("Tandfloss", "rol 50m", 0.33, 2.20, "1 rol per 3 maanden"),
        ("SPF50 dagcrème gezicht", "tube 50ml", 0.5, 12.00, "La Roche-Posay Anthelios / Altijd"),
        ("Kristaldeodorant (aluin)", "stuk 120g", 0.08, 8.00, "1 stuk per jaar"),
        ("Lipverzorging SPF", "stick", 0.17, 3.00, "1 stick per 6 maanden"),
        ("Wattenstaafjes (herbruikbaar)", "set — jaarlijks", 0.08, 10.00, "LastSwab of equiv."),
        ("D-panthenol bodylotion (droge huid)", "fles 200ml", 0.25, 8.00, "Optioneel bij droge huid"),
    ]

    r = 4
    totaal = 0
    for item, eenheid, qty, prijs, toel in items:
        maandkost = round(qty * prijs, 2)
        totaal += maandkost
        bg = C["amber"] if "Optioneel" in toel else "alt"
        drow(ws, r, [item, eenheid, qty, prijs, maandkost, toel],
             bg=bg,
             fmts=[None, None, None, EUR, EUR, None])
        r += 1

    trow(ws, r, ["TOTAAL (excl. optionele items)", "", "", "",
                 round(totaal - 2.00, 2), ""])
    ws.cell(r, 5).number_format = EUR

def onderhoud(wb):
    ws = mk_ws(wb, "Onderhoud", "1A2B4A")
    set_cols(ws, [28, 34, 16, 12, 24, 14])
    title_block(ws, 1, "Onderhoudsblad — Hygiëne")
    hdr(ws, 3, ["Item / Routine", "Taak", "Frequentie", "Duur (min)", "Materialen", "Jaarkosten (€)"])

    taken = [
        ("Ochtend routine", "Douche + tanden + gezicht + deodorant", "Dagelijks", 15, "Zeep, tandpasta, SPF50", 0),
        ("Avond routine", "Tanden + huidreiniging", "Dagelijks", 5, "Tandpasta, make-upremover", 0),
        ("Scheren", "Nat scheren met DE-mes", "3-4× / week", 8, "Scheercrème, mesje", 0),
        ("Nagels knippen", "Vingers + tenen", "2× / maand", 10, "Nagelset", 0),
        ("Kapperbezoek", "Knipbeurt (norm)", "1× / 6 weken", 45, "Kosten in budget", 120),
        ("Elektrische tandenborstel", "Opzetborstel vervangen", "1× / 3 maanden", 2, "Opzetborstel", 32),
        ("Handdoeken wassen", "Was 60°", "Wekelijks", 10, "Wasmiddel", 20),
        ("Douche + wastafel reinigen", "Kalkaanslag, schimmelpreventie", "Wekelijks", 10, "Kalk/schimmelreiniger", 15),
        ("Scheermeshouder + spiegel", "Reinigen met azijn", "Maandelijks", 5, "Wittewijnazijn", 3),
    ]

    r = 4
    for groep, taak, freq, duur, mat, jk in taken:
        drow(ws, r, [groep, taak, freq, duur, mat, jk],
             fmts=[None, None, None, None, None, EUR0])
        r += 1

    trow(ws, r, ["TOTAAL JAARKOSTEN", "", "", "", "", sum(t[5] for t in taken)])
    ws.cell(r, 6).number_format = EUR0

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Hygiëne (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Hygiënemateriaal (11 items)", 234, "", "Eenmalige aankoopkost"])
    trow(ws, 6, ["Totaal investering", 234, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Afschrijving materiaal", 4.12, "Op basis van levensduur per item"),
        ("Verbruiksgoederen (norm)", 23.10, "Excl. optionele items"),
        ("Kapperbezoek", 10.00, "€60 per 6 weken = €120/jaar"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+3, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Ochtend routine (dagelijks)", 450, ""),
        ("Avond routine (dagelijks)", 150, ""),
        ("Scheren (3-4×/week)", 120, ""),
        ("Overig (nagels, kappers prep)", 30, ""),
    ]
    r2 = r + 4
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Hygiëne")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Hygiëne", "Wonen / Shelter", "Badkamer met douche, wastafel, spiegel en warm water vereist."),
        ("Hygiëne", "Wonen / Shelter", "Stopcontact in badkamer voor tandenborstel oplader."),
        ("Hygiëne", "Kleding", "Handdoeken gedeeld met wasroutine kleding (60°)."),
        ("Hygiëne", "Eten", "Tanden poetsen kan aan afwasbak — deelt wateraansluiting."),
        ("Wonen", "Hygiëne", "Watertemperatuur boiler min. 60°C voor hygiënisch warm water."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    verbruik(wb)
    onderhoud(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
