#!/usr/bin/env python3
"""
Lightfoot — export_csv_mirrors.py
Exporteer alle tabbladen van alle systeembestanden als CSV-snapshots.
Gebruik: python3 export_csv_mirrors.py [YYYY-MM-DD] [HHhMM]

Mapstructuur output:
  lightfoot/spreadsheets/<functie>/csv_mirror_snapshots/<datum_tijd>/
"""

import os, sys, csv
from datetime import datetime
from openpyxl import load_workbook

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

BESTANDEN = {
    "kleding_systeem":   "kleding_systeem.xlsx",
    "slaapsysteem":      "slaapsysteem.xlsx",
    "eetsysteem":        "eetsysteem.xlsx",
    "hygienesysteem":    "hygienesysteem.xlsx",
    "werksysteem":       "werksysteem.xlsx",
    "beweging_sport":    "beweging_sport.xlsx",
    "mobiliteit":        "mobiliteit.xlsx",
    "wonen_shelter":     "wonen_shelter.xlsx",
    "leven_inventaris":  "leven_inventaris.xlsx",
}

def export_workbook(functie, bestand, snapshot_dir):
    path = os.path.join(BASE, bestand)
    if not os.path.exists(path):
        print(f"  NIET GEVONDEN: {path}")
        return

    wb = load_workbook(path, data_only=True)
    os.makedirs(snapshot_dir, exist_ok=True)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        safe_name = ws_name.replace("/", "-").replace(" ", "_").lower()
        csv_path = os.path.join(snapshot_dir, f"{safe_name}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([v if v is not None else "" for v in row])
        print(f"    → {os.path.relpath(csv_path, BASE)}")

def main():
    now = datetime.now()
    datum = sys.argv[1] if len(sys.argv) > 1 else now.strftime("%Y-%m-%d")
    tijdlabel = sys.argv[2] if len(sys.argv) > 2 else now.strftime("%Hh%M")
    snapshot_prefix = f"{datum}_{tijdlabel}"

    print(f"\nLightfoot CSV-export — snapshot {snapshot_prefix}")
    print("=" * 60)

    for functie, bestand in BESTANDEN.items():
        snap_dir = os.path.join(BASE, functie, "csv_mirror_snapshots", snapshot_prefix)
        print(f"\n[{functie}]")
        export_workbook(functie, bestand, snap_dir)

    print(f"\nKlaar. Snapshots in: lightfoot/spreadsheets/<functie>/csv_mirror_snapshots/{snapshot_prefix}/")

if __name__ == "__main__":
    main()
