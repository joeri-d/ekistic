"""Lightfoot — mobiliteit.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/mobiliteit.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "MOBILITEIT — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# MOBILITEIT",
        "",
        "## Ontwerpdoel",
        "Maximale verplaatsingsvrijheid met minimale eigendomslast. Geen persoonlijk voertuig",
        "tenzij woonlocatie dit functioneel vereist (ruraal, geen OV).",
        "",
        "## Standaardscenario: stedelijk / randstedelijk Vlaanderen",
        "→ Primair: kwaliteitsfiets voor dagelijkse verplaatsingen ≤ 15 km.",
        "→ Secundair: De Lijn / NMBS voor langere trajecten en slecht weer.",
        "→ Tertiair: Cambio autodelen voor uitzonderlijke ritten (meubels, periferiebezoek).",
        "→ Geen eigen wagen: besparing €400-600/maand na alle kosten.",
        "",
        "## Standaardfiets",
        "→ Model: Gazelle Medeo T10 HMB of Batavus Dinsdag+ — stevige stadsfiets,",
        "   naaldlager, 8-10 versnellingen, geïntegreerde verlichting, dynamo.",
        "→ Accessoires: bel (wettelijk), slot (AXA Defender of Abus Granit), fietstassen.",
        "→ Regenkledij fietsen: regenjasje (deelt met kledingfunctie) + overschoenen optioneel.",
        "",
        "## Openbaar vervoer",
        "→ De Lijn: maandabonnement zone 1-2 of heel Vlaanderen afhankelijk van traject.",
        "→ NMBS: losse tickets of Go Pass 10 voor occasioneel treingebruik.",
        "→ Standaard: De Lijn maandabonnement Gent/Antwerpen/Leuven zone ≈ €49-€54/maand.",
        "",
        "## Uitzondering: persoonlijk voertuig (ruraal)",
        "→ Enkel indien OV-aanbod < 2 verbindingen/uur op werktraject.",
        "→ Standaard auto: Toyota Yaris Hybrid 2e hands (€12.000-15.000).",
        "→ Totale maandkost auto (TCO): geschat €450-600/maand (afschrijving + brandstof +",
        "   verzekering + onderhoud).",
        "→ Niet opgenomen in basismodel — budgetteer apart bij nood.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [28, 34, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Mobiliteit (standaard stadsscenario)")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        ("Stadsfiets", "8 versnellingen, dynamo, spatborden, drager", "Gazelle Medeo T10 HMB", 1, 1099, 120),
        ("Fietsslot", "ART 4-ster, beugel + ketting", "Abus Granit 460/150", 1, 75, 60),
        ("Fietstassen (set)", "Waterproof, 40L totaal, bagagedrager", "Ortlieb Back-Roller City", 1, 130, 120),
        ("Helm", "MIPS, CE EN 1078", "Giro Register MIPS", 1, 55, 60),
        ("Fietshandschoenen", "Touchscreen, winter + zomer set", "Decathlon", 2, 12, 24),
        ("Binnenband (reserve)", "28\", Presta ventiel", "Schwalbe", 2, 6, 12),
        ("Lichtset (backup)", "USB oplaadbaar voor/achter", "Knog Blinder", 1, 35, 36),
        ("Regenoverschoenen", "Neopreen, maat universeel", "Decathlon", 1, 15, 48),
        ("Cambio lidmaatschap", "Basisregistratie (geen maandkost)", "Cambio", 1, 10, 120),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def verbruik(wb):
    ws = mk_ws(wb, "Verbruiksgoederen", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 20])
    title_block(ws, 1, "Verbruiksgoederen — Mobiliteit")
    hdr(ws, 3, ["Item", "Eenheid", "Qty / jaar", "Prijs/eenheid (€)",
                "Jaarkost (€)", "Toelichting"])

    items = [
        ("Binnenband vervanging", "stuk", 2, 8.00, "Pech onderweg"),
        ("Fietsband (buitenband)", "stuk", 0.5, 35.00, "1× per 2 jaar"),
        ("Fietsonderhoud (professioneel)", "beurt", 1, 45.00, "Jaarlijkse servicebeurt"),
        ("Ketting + smeermiddel", "set", 1, 18.00, "Ketting 1×/jaar vervangen"),
        ("Remblokken", "set", 1, 12.00, "1× per jaar bij intensief gebruik"),
        ("De Lijn maandabonnement (zone 2)", "maand", 12, 54.00, "Standaard stadszone"),
        ("NMBS Go Pass 10", "pas", 2, 53.00, "Occasioneel treingebruik"),
        ("Cambio gebruik (geschat)", "maand", 12, 20.00, "Gemiddeld 1-2 ritten/maand"),
    ]

    r = 4
    totaal_jaar = 0
    for item, eenheid, qty, prijs, toel in items:
        jaar = round(qty * prijs, 2)
        totaal_jaar += jaar
        drow(ws, r, [item, eenheid, qty, prijs, jaar, toel],
             fmts=[None, None, None, EUR, EUR0, None])
        r += 1

    trow(ws, r, ["TOTAAL JAAR", "", "", "", totaal_jaar, ""])
    ws.cell(r, 5).number_format = EUR0
    r += 1
    trow(ws, r, ["TOTAAL / MAAND", "", "", "", round(totaal_jaar / 12, 2), ""])
    ws.cell(r, 5).number_format = EUR

def onderhoud(wb):
    ws = mk_ws(wb, "Onderhoud", "1A2B4A")
    set_cols(ws, [28, 34, 16, 12, 24, 14])
    title_block(ws, 1, "Onderhoudsblad — Fiets")
    hdr(ws, 3, ["Item", "Taak", "Frequentie", "Duur (min)", "Materialen", "Jaarkosten (€)"])

    taken = [
        ("Ketting", "Reinigen + smeren", "Maandelijks", 10, "Kettingontetter, smeermiddel", 8),
        ("Banden", "Luchtdruk controleren", "Wekelijks", 3, "Fietspomp", 0),
        ("Remmen", "Spanning + slijtage controleren", "Maandelijks", 5, "-", 0),
        ("Verlichting", "Dynamo check, backup opladen", "Maandelijks", 3, "-", 0),
        ("Volledige fiets", "Professionele jaarlijkse beurt", "1× / jaar", 0, "Vakfietsenzaak", 45),
        ("Ketting vervanging", "Meten + vervangen", "1× / jaar", 15, "Ketting + gereedschap", 18),
        ("Slot + sleutel", "Smeren slotkast", "2× / jaar", 3, "Siliconenspray", 2),
        ("Fietstassen", "Reinigen + waterdichtheid controleren", "Seizoensmatig", 10, "Water + doek", 0),
    ]

    r = 4
    for groep, taak, freq, duur, mat, jk in taken:
        drow(ws, r, [groep, taak, freq, duur, mat, jk],
             fmts=[None, None, None, None, None, EUR0])
        r += 1

    trow(ws, r, ["TOTAAL JAARKOSTEN", "", "", "", "", sum(t[5] for t in taken)])
    ws.cell(r, 6).number_format = EUR0

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Mobiliteit (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Fiets + accessoires (9 items)", 1449, "", "Eenmalige aankoopkost"])
    trow(ws, 6, ["Totaal investering", 1449, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Afschrijving fiets + accessoires", 13.32, "Op basis van levensduur"),
        ("De Lijn maandabonnement zone 2", 54.00, "Standaard stadszone"),
        ("NMBS occasioneel (Go Pass 10)", 8.83, "2 passen per jaar"),
        ("Cambio gebruik (geschat)", 20.00, "1-2 ritten per maand"),
        ("Fietsonderhoud + verbruik", 6.08, "Banden, ketting, service"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "UITZONDERING: Eigen wagen (ruraal)").font = fB(col=C["navy_mid"])
    hdr(ws, r+3, ["Voertuig", "Prijs 2e hands (€)", "TCO / maand (€)", ""],
        bg=C["amber"])
    drow(ws, r+4, ["Toyota Yaris Hybrid", 13500, 520, "Afschr + verz + brandstof + onderhoud"],
         bg=C["amber"])

    ws.cell(r+6, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+7, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Fietsonderhoud eigen beheer", 30, ""),
        ("OV-planning en wachttijden", 60, ""),
    ]
    r2 = r + 8
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Mobiliteit")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Mobiliteit", "Wonen / Shelter", "Fietsenstalling: minstens 1 veilige stalling in of bij woning vereist."),
        ("Mobiliteit", "Werk", "Thuiswerk reduceert pendelbehoefte — mobiliteitsbudget naar beneden."),
        ("Mobiliteit", "Beweging", "Dagelijks fietsen telt mee als bewegingsactiviteit."),
        ("Mobiliteit", "Wonen / Shelter", "Locatiekeuze woning bepaalt OV-nood en fietsafstand werk."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    verbruik(wb)
    onderhoud(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
