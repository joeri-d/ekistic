"""Lightfoot — wonen_shelter.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/wonen_shelter.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "WONEN & SHELTER — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# WONEN & SHELTER",
        "",
        "## Ontwerpdoel",
        "Minimale m² die alle leeffuncties volledig ondersteunen — geen overtollige ruimte,",
        "geen overtollig meubilair. Elke extra m² = meer poets, meer huur, meer onderhoud.",
        "",
        "## Standaardnorm",
        "→ Type: studio of 1-slaapkamerappartement, 28-40 m².",
        "→ Minimale eisen: badkamer met douche, keukenhoek met koelkast, 1 slaapzone,",
        "   1 werkzone, voldoende lichtinval, verwarmingsinstallatie.",
        "→ Locatie: wandelbaar tot supermarkt (≤ 500m) + OV-halte (≤ 300m) + fietsenstalling.",
        "→ Norm huurprijs 2026: €700-850/maand voor studio in middelgrote Vlaamse stad.",
        "→ Energieprestatie: EPC-label min. C — vermijd slechte isolatie (hoge energiekost).",
        "",
        "## Inrichting (minimalistisch)",
        "→ Slaapzone: bed (zie slaapsysteem) + 1 kledingrail of kleine kledingkast.",
        "→ Werkzone: bureau + ergonomische stoel (zie werksysteem) + 1 boekenplank.",
        "→ Leefzone: 1 zitmeubel (fauteuil of 2-zits sofa) + 1 bijzettafel.",
        "→ Keuken: koelkast + inductieplaaat + spoelbak — gedeeld met leeffunctie eten.",
        "→ Opbergruimte: 1 multifunctionele kast (kleding + linnen + diverse).",
        "→ Geen eetkamertafel apart — bureau dient als eet- en werkblad.",
        "",
        "## Reinigingslast per m²",
        "→ 30 m² = ca. 45-60 min wekelijks (stofzuigen, dweilen, badkamer, keuken).",
        "→ Elke extra 10 m² = +15 min/week = +3 uur/jaar onderhoudstijd.",
        "→ Minimale m² is dus ook een tijdsinvestering — geen sentimentele keuze.",
        "",
        "## Utilities standaard",
        "→ Elektriciteit: geschat 150-200 kWh/maand voor studio → €30-40/maand.",
        "→ Water: €15-20/maand (inclusief afvalwater).",
        "→ Verwarming: warmtepomp of stadsverwarming prioritair — gas als norm in bestaand park.",
        "→ Internet: opgenomen in werksysteem.",
        "→ Woonverzekering: brand + diefstal + aansprakelijkheid → €15-20/maand.",
        "",
        "## Uitzondering",
        "→ Eigen woning (hypotheek): apart scenario, niet opgenomen in basismodel.",
        "→ Kotgemeenschap of gedeelde woning: kan maandkost halveren maar compliciteert",
        "   leeffuncties (keuken, hygiëne, geluid) — bewuste keuze, niet standaard.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [28, 34, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Inboedel (excl. bed en bureau, zie systeemfiles)")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        ("Kledingrail + lades set", "Staal, 100cm + 2 laden", "IKEA Aurdal", 1, 120, 120),
        ("Boekenplank", "180×80cm, 5 niveaus, wit", "IKEA Billy", 1, 60, 120),
        ("Fauteuil / 2-zits sofa", "Stof, lichtgrijs, compact", "IKEA Friheten of Koarp", 1, 250, 120),
        ("Bijzettafel", "Ronde bijzettafel 50cm, bamboe", "IKEA Svalsta", 1, 30, 120),
        ("Vloerlamp", "LED, dimbaar, E27, zwart", "IKEA Ranarp", 1, 40, 120),
        ("Bureaulamp", "LED, dimbaar, USB-C, architectenlamp", "IKEA Forsa", 1, 30, 120),
        ("Multifunctionele kast", "180×80×40cm, 4-deurs", "IKEA Brimnes", 1, 180, 120),
        ("Badkamerkast klein", "Wandmontage, spiegel", "IKEA Lillången", 1, 60, 120),
        ("Douchegordijn + rail + ringen", "Waterbestendig, wit", "IKEA Bjärsen", 1, 25, 36),
        ("Vuilnisbakken set", "20L + 5L gescheiden afval", "Brabantia Sort & Go", 1, 40, 120),
        ("Basisgereedschap set", "Hamer, schroevendraaiers, waterpas, boor", "Bahco / Stanley kit", 1, 55, 240),
        ("Stofzuiger", "Draadloos, HEPA, 30 min auton.", "Dyson V8 Slim of Miele Triflex", 1, 199, 84),
        ("Mopset (vloer)", "Microvezel, 360° wasbaar", "Leifheit Twist", 1, 35, 48),
        ("Rookmelder + CO-melder", "10 jaar batterij, wettelijk vereist", "FireAngel", 2, 20, 120),
        ("Verdeelstekker (surge protected)", "6-voudig, schakelbaar", "Brennenstuhl", 2, 20, 120),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL INBOEDEL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def verbruik(wb):
    ws = mk_ws(wb, "Verbruiksgoederen", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 20])
    title_block(ws, 1, "Verbruiksgoederen — Woning (poetsen + onderhoud)")
    hdr(ws, 3, ["Item", "Eenheid", "Qty / jaar", "Prijs/eenheid (€)",
                "Jaarkost (€)", "Toelichting"])

    items = [
        ("Allesreiniger (geconcentreerd)", "fles 1L", 3, 4.00, "Vloer, aanrecht, badkamer"),
        ("Toiletreiniging gel", "fles 750ml", 4, 2.50, "Wekelijkse toilet"),
        ("Schimmelverwijderaar (badkamer)", "spuitbus 500ml", 2, 5.00, "Kwartaalonderhoud"),
        ("Kalkreiniger (badkamer + keuken)", "fles 750ml", 3, 3.50, "Maandelijks kalkaanslag"),
        ("Microvezel doeken set", "set 5 stuks", 2, 6.00, "Vernieuwen 2× per jaar"),
        ("Vuilniszakken (PMD + restafval)", "rol 20st", 8, 3.00, "Wekelijks gebruik"),
        ("Glas- en raamreiniger", "spuitbus 750ml", 2, 3.00, "Kwartaalramen"),
        ("Boenwas / vloeronderhoud", "fles 750ml", 1, 8.00, "Jaarlijks parket/laminaat"),
        ("Reservelamp LED E27", "stuk", 2, 4.00, "Vervanging bij uitval"),
    ]

    r = 4
    totaal_jaar = 0
    for item, eenheid, qty, prijs, toel in items:
        jaar = round(qty * prijs, 2)
        totaal_jaar += jaar
        drow(ws, r, [item, eenheid, qty, prijs, jaar, toel],
             fmts=[None, None, None, EUR, EUR0, None])
        r += 1

    trow(ws, r, ["TOTAAL JAAR", "", "", "", totaal_jaar, ""])
    ws.cell(r, 5).number_format = EUR0
    r += 1
    trow(ws, r, ["TOTAAL / MAAND", "", "", "", round(totaal_jaar / 12, 2), ""])
    ws.cell(r, 5).number_format = EUR

def reinigingsritme(wb):
    ws = mk_ws(wb, "Reinigingsritme", "2D4A7A")
    set_cols(ws, [24, 40, 16, 12, 20])
    title_block(ws, 1, "Reinigingsritme — Studio 30 m²")
    hdr(ws, 3, ["Zone", "Taak", "Frequentie", "Duur (min)", "Materialen"])

    taken = [
        ("Vloer volledig", "Stofzuigen", "Wekelijks", 15, "Stofzuiger"),
        ("Vloer volledig", "Dweilen / moppenf", "Wekelijks", 15, "Mop + allesreiniger"),
        ("Badkamer", "Toilet + douche + wastafel", "Wekelijks", 20, "Toilet gel, kalkreiniger"),
        ("Keuken", "Aanrecht + spoelbak + plaat", "Wekelijks", 10, "Allesreiniger"),
        ("Ramen + spiegels", "Sprayen + afnemen", "Maandelijks", 15, "Raamreiniger"),
        ("Koelkast", "Uitruimen + reinigen", "Maandelijks", 15, "Allesreiniger"),
        ("Stofzuiger leegmaken / filter", "Leegmaken + filter reinigen", "Maandelijks", 10, "-"),
        ("Volledige woning dieptereiniging", "Alle hoeken, kasten, gordijnen", "Seizoensmatig", 120, "Volledig pakket"),
        ("Schimmelpreventie badkamer", "Ventileren + anti-schimmel spray", "Kwartaal", 10, "Schimmelspray"),
    ]

    r = 4
    for zone, taak, freq, duur, mat in taken:
        drow(ws, r, [zone, taak, freq, duur, mat])
        r += 1

    ws.cell(r+1, 1, f"Geschatte wekelijkse tijd: 60 min — maandelijkse tijdkost: 4-5 uur").font = fI()

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Wonen (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING INBOEDEL").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Inboedel volledig (15 items)", 1204, "", "Excl. bed, bureau, stoel (zie systeemfiles)"])
    trow(ws, 6, ["Totaal investering inboedel", 1204, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Huur (norm studio 30-35 m²)", 750.00, "Mediane huur middelgrote Vlaamse stad 2026"),
        ("Elektriciteit (150-200 kWh)", 35.00, "EPC C, warmtepomp of gas CV"),
        ("Water + afvalwater", 18.00, "Gemiddeld verbruik 1 persoon"),
        ("Woonverzekering (brand + diefstal)", 17.00, "BA + brand + inbraak"),
        ("Gemeenschappelijke kosten", 40.00, "Liften, gang, tuinonderhoud indien van toepassing"),
        ("Afschrijving inboedel", 8.92, "Op basis van levensduur per item"),
        ("Verbruiksgoederen (poetsen)", 5.25, "Reinigingsmiddelen"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+3, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Wekelijkse reiniging (4× 60 min)", 240, ""),
        ("Maandelijkse taken (ramen, koelkast)", 30, ""),
        ("Seizoensreiniging + beheer", 30, ""),
    ]
    r2 = r + 4
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Wonen")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Wonen", "Alle functies", "Shelter is de infrastructuurlaag voor alle andere leeffuncties."),
        ("Wonen", "Slaap", "Slaapzone: min. 10 m², temperatuur regelbaar 16-18 °C."),
        ("Wonen", "Eten", "Keukenhoek: koelkast + 60 cm werkblad + spoelbak + stopcontact vereist."),
        ("Wonen", "Hygiëne", "Badkamer: douche + wastafel + warm water + ventilatie vereist."),
        ("Wonen", "Werk", "Werkzone: stil, daglijcht, stopcontact, min. 6 m² vrije ruimte."),
        ("Wonen", "Beweging", "Bewegingsvloer: 2×2 m vrij voor mat + kettlebell vereist."),
        ("Wonen", "Mobiliteit", "Fietsenstalling: veilige, overdekte stalling bij woning."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    verbruik(wb)
    reinigingsritme(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
