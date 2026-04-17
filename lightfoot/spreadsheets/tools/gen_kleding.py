"""Lightfoot — kleding_systeem.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/kleding_systeem.xlsx"

EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "KLEDING — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# KLEDING",
        "",
        "## Ontwerpdoel",
        "Een capsulegarderobe die maximale functionaliteit biedt met minimale items, nul modedruk,",
        "nul keuzestress en een lage totale eigendomskost (TCO) over de volledige levensduur.",
        "",
        "## Kernprincipes",
        "→ Palet: navy / charcoal / wit / zwart — volledig onderling combineerbaar.",
        "→ Materiaal: merino wol als standaard voor tops en sokken (temperatuurregulerend,",
        "   geurresistent, machinewasbaar op 30°, lange levensduur).",
        "→ Technische stoffen (stretch, DWR) voor broeken en buitenlaag.",
        "→ Geen modieuze coupe — tijdloze, slanke snit die decennia meegaat.",
        "→ Repareer vóór vervanging. Kleine herstelling = 15 min, verlengt levensduur met jaren.",
        "",
        "## Standaard merken",
        "→ Uniqlo: merino basics, Oxford-hemd — uitstekende prijs-kwaliteitsverhouding.",
        "→ Decathlon (Forclaz/Trek): technische broeken, sokken, buitenlaag — goedkoop, functioneel.",
        "→ New Balance 990 of HOKA Clifton: enige schoenkeuze voor dagelijks gebruik.",
        "→ Salomon X Ultra 4 Mid GTX: waterdichte wandelschoen voor regenperiodes en terrein.",
        "",
        "## Uitzonderingen (niet opgenomen in standaard)",
        "→ Formeel pak: alleen nodig bij begrafenis/bruiloft — huren of éénmalige aankoop buiten model.",
        "→ Sportkleding: opgenomen in beweging_sport.xlsx.",
        "→ Badjas: opgenomen in hygienesysteem.xlsx.",
        "",
        "## Onderhoud",
        "→ Was alles op 30°, binnenstebuiten, in waszak voor merinowol.",
        "→ Droger: nooit voor merino, bij voorkeur ophangen.",
        "→ Strijken: niet nodig bij juiste was- en droogmethode.",
        "→ Jassen DWR-behandeling: 2× per jaar met Nikwax TX.Direct.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [28, 32, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Minimale kledinggarderobe")
    ws.row_dimensions[3].height = 14
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        # NORM-rij = wit/grijs afwisselend | UITZONDERING = amber
        # (item, spec, merk, qty, prijs, levensduur_mnd, bg)
        ("Merino T-shirt", "150 g/m², V-hals, navy / charcoal / wit", "Uniqlo Supima Cotton", 4, 25, 48, "alt"),
        ("Merino lange mouw", "150 g/m², crew neck", "Uniqlo Extra Fine Merino", 2, 35, 48, "alt"),
        ("Merino pullover", "200 g/m², crew neck, navy", "Uniqlo Fine Knit", 1, 50, 60, "alt"),
        ("Softshell jack", "Waterproof, DWR, zwart", "Decathlon Forclaz 900", 1, 95, 84, "alt"),
        ("Technische broek", "Stretch, slim, charcoal", "Decathlon NH500", 2, 60, 60, "alt"),
        ("Boxer short", "Bamboo/merino, zwart", "Bamboo Basics", 4, 12, 24, "alt"),
        ("Merino sokken", "Mid-weight, charcoal", "Decathlon Kiprun", 4, 8, 24, "alt"),
        ("Sneaker dagelijks", "All-weather, wit/zwart", "New Balance 990 v6", 1, 160, 36, "alt"),
        ("Waterdichte wandelschoen", "Mid-cut GTX", "Salomon X Ultra 4 Mid GTX", 1, 160, 60, "alt"),
        ("Oxford hemd", "Wit, slim fit, formele context", "Uniqlo Oxford", 1, 35, 72, "alt"),
        ("Blazer", "Navy, slim fit, tijdloos", "COS / Zara Man", 1, 120, 84, "alt"),
        ("Thuisbroek", "Losse trainingsbroek, binnenkleding", "Decathlon", 1, 20, 60, "alt"),
        ("Riem", "Zwart leder, zilver gesp", "n/a", 1, 30, 120, "alt"),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd, bg in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             bg=bg,
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def verbruik(wb):
    ws = mk_ws(wb, "Verbruiksgoederen", "2D4A7A")
    set_cols(ws, [28, 14, 16, 14, 14, 16])
    title_block(ws, 1, "Verbruiksgoederen — Kleding")
    hdr(ws, 3, ["Item", "Eenheid", "Qty / jaar", "Prijs/eenheid (€)",
                "Jaarkost (€)", "Maandkost (€)"])

    items = [
        ("Wascapsules kleur/donker", "doos 35st", 2, 15),
        ("Merino wasmiddel (Woolite/Sport)", "fles 1L", 2, 11),
        ("Vlekkenverwijderaar (Vanish)", "stick 75g", 1, 5),
        ("Waszak merino (Guppyfriend)", "stuk", 1, 15),
        ("Nikwax TX.Direct (DWR jack)", "fles 300ml", 1, 14),
        ("Impregneer spray schoenen", "bus 300ml", 1, 9),
        ("Naald + draad herstellingsset", "set", 1, 5),
    ]

    r = 4
    totaal_jaar = 0
    for item, eenheid, qty, prijs in items:
        jaar = qty * prijs
        maand = round(jaar / 12, 2)
        totaal_jaar += jaar
        drow(ws, r, [item, eenheid, qty, prijs, jaar, maand],
             fmts=[None, None, None, EUR, EUR0, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", totaal_jaar, round(totaal_jaar / 12, 2)])
    ws.cell(r, 5).number_format = EUR0
    ws.cell(r, 6).number_format = EUR

def onderhoud(wb):
    ws = mk_ws(wb, "Onderhoud", "1A2B4A")
    set_cols(ws, [28, 32, 16, 12, 24, 14])
    title_block(ws, 1, "Onderhoudsblad — Kleding")
    hdr(ws, 3, ["Item / Groep", "Taak", "Frequentie", "Duur (min)", "Materialen", "Jaarkosten (€)"])

    taken = [
        ("T-shirts, boxers, sokken", "Was 30°, binnenstebuiten, ophangen", "Wekelijks", 10, "Wascapsule", 15),
        ("Merino pullover / lange mouw", "Was 30° delicaat in waszak, plat drogen", "2× / maand", 15, "Woolite, waszak", 11),
        ("Broeken (technisch)", "Was 30°, ophangen", "2× / maand", 10, "Wascapsule", 8),
        ("Sneakers", "Reinigen met doek, impregneren", "Maandelijks", 10, "Impregneer spray", 9),
        ("Waterdichte wandelschoen", "Reinigen + DWR herlaaien", "2× / jaar", 20, "Nikwax TX.Direct", 14),
        ("Softshell jack", "Was 40° + DWR herlaaien", "2× / jaar", 20, "Nikwax TX.Direct", 14),
        ("Volledige inventaris", "Seizoenscheck: slijtage, herstellen, vervangen", "2× / jaar", 30, "Naald + draad", 5),
        ("Riem", "Afnemen met vochtige doek, conditioneren", "1× / jaar", 5, "Ledercondicioner", 5),
    ]

    r = 4
    for groep, taak, freq, duur, mat, jk in taken:
        drow(ws, r, [groep, taak, freq, duur, mat, jk],
             fmts=[None, None, None, None, None, EUR0])
        r += 1

    trow(ws, r, ["TOTAAL JAARKOSTEN", "", "", "", "", sum(t[5] for t in taken)])
    ws.cell(r, 6).number_format = EUR0

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Kleding (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "Toelichting", ""])
    drow(ws, 5, ["Totale garderobe (13 items)", 955, "Eenmalige aankoopkost", ""])
    drow(ws, 6, ["Wisselgarderobe buffer (10%)", 96, "Reserve voor onverwachte vervanging", ""])
    trow(ws, 7, ["Totaal investering", 1051, "", ""])
    ws.cell(7, 2).number_format = EUR0

    ws.cell(9, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 10, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Afschrijving garderobe", 19.13, "Op basis van levensduur per item"),
        ("Verbruiksgoederen (was/onderhoud)", 6.42, "Wasmiddel, sprays, vlekken"),
        ("Herstellingen (gemiddeld)", 2.00, "Naad, knoop, zool — geschat"),
    ]
    r = 11
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+3, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Wassen + drogen + ophangen", 80, ""),
        ("Seizoenscheck + herstelling", 10, ""),
        ("Schoen- en jackonderhoud", 10, ""),
    ]
    r2 = r + 4
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Kleding")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])

    afh = [
        ("Kleding", "Wonen / Shelter", "Opbergruimte: 1 kast of kledingrail + lades vereist."),
        ("Kleding", "Hygiëne", "Wasgelegenheid: wasmachine of wasplaats vereist."),
        ("Kleding", "Beweging / Sport", "Sportkleding apart gebudgetteerd in beweging_sport."),
        ("Kleding", "Werk", "Geen formeel uniform vereist in standaardmodel."),
        ("Hygiëne", "Kleding", "Schone kleding vereist regelmatig wassen — ritme afgestemd."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    verbruik(wb)
    onderhoud(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
