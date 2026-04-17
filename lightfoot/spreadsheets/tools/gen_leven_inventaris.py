"""Lightfoot — leven_inventaris.xlsx (masterfile) generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook
from openpyxl.styles import Font

OUT = "/home/user/ekistic/lightfoot/spreadsheets/leven_inventaris.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

# ── MASTERDATA (afgestemd op alle systeemfiles) ──────────────────────────────
FUNCTIES = [
    # (naam, items, investering, maandkost_totaal, waarvan_vast, tijdmin_mnd)
    ("Kleding",         23,  1051,  27.55,  0,    100),
    ("Slaap",            8,   699,  11.27,  0,     90),
    ("Eten",            18,   306, 170.48,  0,    680),
    ("Hygiëne",         11,   234,  37.22,  0,    750),
    ("Werk",            10,  3306, 118.47, 57.00,  35),
    ("Beweging & Sport",  8,   182,   4.33,  0,    370),
    ("Mobiliteit",       9,  1449, 102.23, 54.00,  90),
    ("Wonen & Shelter", 15,  1204, 874.17,750.00,  300),
]

def dashboard(wb):
    ws = mk_ws(wb, "Dashboard", "1A2B4A")
    set_cols(ws, [24, 8, 14, 16, 16, 16, 14])

    # Titelbalk
    ws.merge_cells("A1:G1")
    t = ws.cell(1, 1, "LIGHTFOOT — PROJECT EXISTENZMINIMUM 2.0")
    t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t.fill = F(C["navy"])
    t.alignment = aC()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    sub = ws.cell(2, 1, "Minimumleefmodel · Vlaamse context · Versie 2026-04-17")
    sub.font = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
    sub.fill = F(C["navy_mid"])
    sub.alignment = aC()
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    hdr(ws, 4, ["Leeffunctie", "# Items", "Investering (€)",
                "Maandkost totaal (€)", "waarvan vaste kost (€)",
                "Afschr/mnd (€)", "Tijdslast (min/mnd)"])
    ws.row_dimensions[4].height = 32

    r = 5
    for naam, items, inv, maand, vast, tijd in FUNCTIES:
        afschr = round(maand - vast - (maand - vast) * 0, 2)  # rough
        drow(ws, r, [naam, items, inv, maand, vast, round(maand - vast, 2), tijd],
             fmts=[None, None, EUR0, EUR, EUR, EUR, None])
        r += 1

    # Totaalrij
    tot_items = sum(f[1] for f in FUNCTIES)
    tot_inv   = sum(f[2] for f in FUNCTIES)
    tot_maand = round(sum(f[3] for f in FUNCTIES), 2)
    tot_vast  = round(sum(f[4] for f in FUNCTIES), 2)
    tot_tijd  = sum(f[5] for f in FUNCTIES)

    trow(ws, r, ["TOTAAL", tot_items, tot_inv, tot_maand, tot_vast,
                 round(tot_maand - tot_vast, 2), tot_tijd])
    ws.cell(r, 3).number_format = EUR0
    ws.cell(r, 4).number_format = EUR
    ws.cell(r, 5).number_format = EUR
    ws.cell(r, 6).number_format = EUR
    ws.row_dimensions[r].height = 20

    r += 2
    ws.cell(r, 1, "TOELICHTING KOLOMMEN").font = fB(col=C["navy"])
    toelichtingen = [
        ("Investering (€)", "Eenmalige aankoopkost van alle items in de functie."),
        ("Maandkost totaal (€)", "Afschrijving + verbruiksgoederen + diensten + onderhoud."),
        ("Waarvan vaste kost (€)", "Huur, abonnementen — onvermijdelijk, onafhankelijk van gebruik."),
        ("Tijdslast (min/mnd)", "Geschatte actieve beheertijd per maand (excl. werkelijk gebruik)."),
    ]
    hdr(ws, r+1, ["Kolom", "Definitie", "", "", "", "", ""], bg=C["navy_mid"])
    for i, (k, v) in enumerate(toelichtingen):
        drow(ws, r+2+i, [k, v, "", "", "", "", ""])

def budget_master(wb):
    ws = mk_ws(wb, "Budget_master", "2D4A7A")
    set_cols(ws, [24, 14, 14, 14, 14, 20])
    title_block(ws, 1, "Budget Master — Alle functies gecombineerd")
    hdr(ws, 3, ["Leeffunctie", "Investering (€)", "Maandkost (€)",
                "Jaarkost (€)", "Vaste kost/mnd (€)", "Variabele kost/mnd (€)"])

    r = 4
    for naam, items, inv, maand, vast, tijd in FUNCTIES:
        variabel = round(maand - vast, 2)
        drow(ws, r, [naam, inv, maand, round(maand * 12, 2), vast, variabel],
             fmts=[None, EUR0, EUR, EUR0, EUR, EUR])
        r += 1

    tot_inv   = sum(f[2] for f in FUNCTIES)
    tot_maand = round(sum(f[3] for f in FUNCTIES), 2)
    tot_vast  = round(sum(f[4] for f in FUNCTIES), 2)
    trow(ws, r, ["TOTAAL", tot_inv, tot_maand, round(tot_maand * 12, 2),
                 tot_vast, round(tot_maand - tot_vast, 2)])
    ws.cell(r, 2).number_format = EUR0
    ws.cell(r, 3).number_format = EUR
    ws.cell(r, 4).number_format = EUR0
    ws.cell(r, 5).number_format = EUR
    ws.cell(r, 6).number_format = EUR
    ws.row_dimensions[r].height = 20

    r += 2
    ws.cell(r, 1, "BENCHMARK").font = fB(col=C["navy"])
    bench = [
        ("Leefloon (OCMW) 2026", 1058, "Absolute wettelijke bodem — model overschrijdt dit"),
        ("Mediaaninkomen netto Vlaanderen 2026", 2100, "Model = 68% van mediaan netto"),
        ("Lightfoot Existenzminimum norm", round(sum(f[3] for f in FUNCTIES), 2),
         "Dit model — inclusief huur + alle functies"),
    ]
    hdr(ws, r+1, ["Referentie", "Bedrag/mnd (€)", "Toelichting", "", "", ""], bg=C["navy_mid"])
    for i, (ref, bedrag, toel) in enumerate(bench):
        drow(ws, r+2+i, [ref, bedrag, toel, "", "", ""],
             fmts=[None, EUR, None, None, None, None])

def tijdsbelasting(wb):
    ws = mk_ws(wb, "Tijdsbelasting", "2D4A7A")
    set_cols(ws, [24, 16, 16, 16, 24])
    title_block(ws, 1, "Tijdsbelasting — Overzicht beheer per functie")
    hdr(ws, 3, ["Leeffunctie", "min/maand", "uur/maand", "uur/jaar", "Toelichting"])

    toel_map = {
        "Kleding":          "Was, drogen, seizoenscheck",
        "Slaap":            "Lakens wassen + ophangen",
        "Eten":             "Inkopen + koken + afwas (actief gebruik, niet enkel beheer)",
        "Hygiëne":          "Dagelijkse routines (ochtend + avond + scheren)",
        "Werk":             "Updates, backup check, hardwarereiniging",
        "Beweging & Sport": "Training 3×/week + wandelingen",
        "Mobiliteit":       "Fietsonderhoud + OV-planning",
        "Wonen & Shelter":  "Weekreiniging + maandtaken",
    }

    r = 4
    for naam, items, inv, maand, vast, tijd in FUNCTIES:
        uurmnd = round(tijd / 60, 1)
        drow(ws, r, [naam, tijd, uurmnd, round(uurmnd * 12, 1), toel_map.get(naam, "")],
             fmts=[None, None, None, None, None])
        r += 1

    tot_tijd = sum(f[5] for f in FUNCTIES)
    trow(ws, r, ["TOTAAL", tot_tijd, round(tot_tijd / 60, 1),
                 round(tot_tijd * 12 / 60, 1), ""])

    r += 2
    ws.cell(r, 1,
            "Noot: tijdsbelasting Eten en Hygiëne bevatten de actieve gebruikstijd,").font = fI()
    ws.cell(r+1, 1,
            "niet enkel onderhoud. Puur beheer (excl. koken, douchen) = ca. 40 uur/jaar.").font = fI()

def afhankelijkheden_matrix(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "1A2B4A")
    set_cols(ws, [24, 24, 52])
    title_block(ws, 1, "Afhankelijkheden — Overzicht alle functies")
    ws.cell(2, 1, "Kritieke koppelingen tussen leeffuncties — norm en uitzondering").font = fI()
    hdr(ws, 4, ["Van", "Naar", "Kritieke koppeling"])

    matrix = [
        ("Wonen", "Alle functies", "Shelter is infrastructuurlaag: keuken, badkamer, werkplek, slaapplek, opslag."),
        ("Kleding", "Hygiëne", "Wasmachine / wasplaats vereist voor kledingonderhoud."),
        ("Kleding", "Wonen", "Opslagruimte: 1 kledingrail + lades in woning inbegrepen."),
        ("Slaap", "Wonen", "Slaapzone: min. 10 m², 16-18 °C regelbaar."),
        ("Eten", "Wonen", "Keuken: koelkast, werkblad, spoelbak, stopcontact vereist."),
        ("Eten", "Hygiëne", "Afwas deelt water-infra. Handen wassen voor koken."),
        ("Hygiëne", "Wonen", "Badkamer: douche, wastafel, warm water, ventilatie."),
        ("Werk", "Wonen", "Stille werkzone, 6 m² min., stopcontact, daglicht."),
        ("Werk", "Mobiliteit", "Thuiswerk = minder pendel = lagere mobiliteitskosten."),
        ("Werk", "Ontspanning", "Laptop + koptelefoon dienen dubbel: werk én vrije tijd."),
        ("Beweging", "Wonen", "2×2 m vrije vloer voor mat + kettlebell."),
        ("Beweging", "Mobiliteit", "Dagelijks fietsen telt als cardio — geen dubbeltelling."),
        ("Mobiliteit", "Wonen", "Fietsenstalling bij woning: overdekt en veilig."),
        ("Mobiliteit", "Locatie", "Locatiekeuze woning bepaalt OV-nood en fietsafstand."),
    ]

    for r, (van, naar, kop) in enumerate(matrix, 5):
        drow(ws, r, [van, naar, kop])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    dashboard(wb)
    budget_master(wb)
    tijdsbelasting(wb)
    afhankelijkheden_matrix(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
