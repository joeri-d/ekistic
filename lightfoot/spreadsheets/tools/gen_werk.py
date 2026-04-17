"""Lightfoot — werksysteem.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/werksysteem.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "WERK — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# WERKSYSTEEM",
        "",
        "## Ontwerpdoel",
        "Een ergonomisch, betrouwbaar en privacybewust werkstation dat kenniswerk,",
        "communicatie en digitale taken volledig ondersteunt — thuis én mobiel.",
        "",
        "## Standaardhardware",
        "→ Laptop: MacBook Air M3 13\" (16 GB / 512 GB) — 10+ uur batterij, passieve koeling,",
        "   volwassen UNIX-basis, lange softwaresupport, uitstekende herverkoopwaarde.",
        "→ Monitor: LG 27UK850-W 4K USB-C — 1 kabel voor stroom + display + USB-hub.",
        "→ Ergonomisch bureau: FlexiSpot E7 Pro zit-sta (140×70) — primaire investering in gezondheid.",
        "→ Stoel: HAG Capisco 8106 of als budget: IKEA Markus — rug en heupen als prioriteit.",
        "→ Toetsenbord: Apple Magic Keyboard met Touch ID (aansluit op macOS-ecosysteem).",
        "→ Muis: Logitech MX Master 3S — ergonomisch, still-click, USB-C.",
        "→ Hoofdtelefoon: Sony WH-1000XM5 — ANC voor focus, communicatie en vrije tijdsgebruik.",
        "   (meervoudig gebruik: werk + ontspanning = geen extra toestel).",
        "",
        "## Connectiviteit",
        "→ Vast internet: Telenet Fiber 500/50 Mbps — €49/maand (of Proximus Fiber equivalent).",
        "→ VPN: Mullvad — €5/maand, no-log, betaling anoniem mogelijk, WireGuard protocol.",
        "→ Publiek wifi: NOOIT zonder VPN actief.",
        "→ Smartphone: deelt mobiele data en hotspot — geen apart mobiel abonnement voor werk.",
        "",
        "## Energiegebruik",
        "→ MacBook Air M3: max. 30W actief, gemiddeld 10-15W.",
        "→ Monitor LG 27UK850: max. 60W.",
        "→ Werktijdgebruik geschat: 8u/dag × 22 werkdagen = 176u/maand.",
        "→ Geschat maandelijks verbruik: ca. 15 kWh → ~€3/maand.",
        "",
        "## Uitzondering",
        "→ NAS / thuisserver: niet opgenomen in basismodel.",
        "→ Tweede monitor: uitzondering bij specifiek professioneel vereiste.",
        "→ Printer: niet opgenomen — cloud print via bibliotheek/copy shop voor uitzonderlijk gebruik.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [28, 34, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Werkstation")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        ("Laptop", "M3, 16GB RAM, 512GB SSD, 13\"", "MacBook Air M3 13\"", 1, 1499, 60),
        ("Monitor", "27\", 4K, USB-C 96W, IPS", "LG 27UK850-W", 1, 449, 84),
        ("Zit-sta bureau", "140×70 cm, elektrisch, stalen frame", "FlexiSpot E7 Pro", 1, 399, 120),
        ("Bureaustoel", "Ergonomisch, lumbaalsteun", "IKEA Markus", 1, 229, 84),
        ("Toetsenbord", "Compact, Touch ID, Bluetooth", "Apple Magic Keyboard", 1, 109, 60),
        ("Muis", "Ergonomisch, still-click, USB-C", "Logitech MX Master 3S", 1, 99, 60),
        ("Koptelefoon", "ANC, USB-C, draadloos", "Sony WH-1000XM5", 1, 349, 60),
        ("Laptop standaard", "Aluminium, verstelbaar", "Rain Design mStand", 1, 49, 120),
        ("USB-C hub (backup)", "7-in-1, 4K HDMI, SD, USB-A×3", "Anker 555", 1, 45, 60),
        ("Webcam (extern, uitzondering)", "1080p, privacy shutter", "Logitech C920", 1, 79, 60),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        bg = C["amber"] if "uitzondering" in item.lower() else "alt"
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             bg=bg,
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def diensten(wb):
    ws = mk_ws(wb, "Diensten", "2D4A7A")
    set_cols(ws, [30, 14, 14, 14, 14, 20])
    title_block(ws, 1, "Diensten — Maandelijkse abonnementen (werk)")
    hdr(ws, 3, ["Dienst", "Aanbieder", "Type", "Prijs/maand (€)", "Prijs/jaar (€)", "Toelichting"])

    items = [
        ("Vast internet fiber", "Telenet Fiber 500", "Abonnement", 49.00, "Basis werkverbinding"),
        ("VPN", "Mullvad", "Abonnement", 5.00, "Privacy, publiek wifi, anoniem"),
        ("Cloud backup", "iCloud 200GB", "Abonnement", 2.99, "Automatische backup laptop"),
        ("Password manager", "Bitwarden (gratis tier)", "Gratis", 0.00, "Wachtwoordbeheer"),
        ("Videocommunicatie", "Zoom / Teams (gratis tier)", "Gratis", 0.00, "Standaard gratis voldoet"),
    ]

    r = 4
    totaal = 0
    for dienst, aanb, typ, prijs, toel in items:
        jaar = round(prijs * 12, 2)
        totaal += prijs
        drow(ws, r, [dienst, aanb, typ, prijs, jaar, toel],
             fmts=[None, None, None, EUR, EUR0, None])
        r += 1

    trow(ws, r, ["TOTAAL / MAAND", "", "", totaal, round(totaal * 12, 2), ""])
    ws.cell(r, 4).number_format = EUR
    ws.cell(r, 5).number_format = EUR0

def onderhoud(wb):
    ws = mk_ws(wb, "Onderhoud", "1A2B4A")
    set_cols(ws, [28, 34, 16, 12, 24, 14])
    title_block(ws, 1, "Onderhoudsblad — Werkstation")
    hdr(ws, 3, ["Item", "Taak", "Frequentie", "Duur (min)", "Materialen", "Jaarkosten (€)"])

    taken = [
        ("Laptop scherm + toetsenbord", "Reinigen met microvezel + isopropyl", "Wekelijks", 5, "Microvezeldoek, IPA", 5),
        ("Monitor", "Afnemen met schermreiniger", "Wekelijks", 3, "Schermreiniger spray", 8),
        ("Bureau + stoel", "Afnemen met vochtige doek", "Wekelijks", 5, "-", 0),
        ("Laptop", "macOS-updates + backup check", "Maandelijks", 15, "-", 0),
        ("USB-C kabels + hub", "Visuele check, vervanging slijtage", "Jaarlijks", 10, "Reserve kabel €10", 5),
        ("Bureau (sta-mechanisme)", "Smeren + check hoogte-instelling", "Jaarlijks", 10, "Siliconenspray", 5),
        ("Koptelefoon", "Oorkussens reinigen, kabel check", "Maandelijks", 5, "Vochtige doek", 0),
        ("Laptop batterijcalibratie", "Check cyclustelling, indien >80% OK", "Jaarlijks", 10, "-", 0),
    ]

    r = 4
    for groep, taak, freq, duur, mat, jk in taken:
        drow(ws, r, [groep, taak, freq, duur, mat, jk],
             fmts=[None, None, None, None, None, EUR0])
        r += 1

    trow(ws, r, ["TOTAAL JAARKOSTEN", "", "", "", "", sum(t[5] for t in taken)])
    ws.cell(r, 6).number_format = EUR0

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Werk (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Hardware werkstation (10 items)", 3306, "", "Eenmalige aankoopkost"])
    trow(ws, 6, ["Totaal investering", 3306, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Afschrijving hardware", 56.98, "Op basis van levensduur per item"),
        ("Internet (Telenet Fiber 500)", 49.00, "Vaste verbinding thuis"),
        ("VPN (Mullvad)", 5.00, "Privacy en beveiliging"),
        ("Cloud backup (iCloud 200GB)", 2.99, "Automatische backup"),
        ("Elektriciteit werkstation", 3.00, "Geschat 15 kWh/maand × €0.20"),
        ("Onderhoud / verbraik materiaal", 1.50, "Reinigingsmiddelen, kabels"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "TIJDSBELASTING (onderhoud)").font = fB(col=C["navy"])
    hdr(ws, r+3, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Updates, backup check, beveiliging", 20, ""),
        ("Reiniging hardware", 15, ""),
    ]
    r2 = r + 4
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Werk")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Werk", "Wonen / Shelter", "Bureau + stoel zijn geïntegreerd in shelter-functie (ruimte)."),
        ("Werk", "Wonen / Shelter", "Stopcontact met geaarde stroomvoorziening + UPS optioneel."),
        ("Werk", "Mobiliteit", "Thuiswerk reduceert mobiliteitskosten — primaire werkplek = thuis."),
        ("Werk", "Ontspanning (later)", "Laptop + koptelefoon dienen dubbel voor werk en vrije tijd."),
        ("Werk", "Hygiëne", "Zittend werk vereist actieve bewegingspauzes — zie beweging_sport."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    diensten(wb)
    onderhoud(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
