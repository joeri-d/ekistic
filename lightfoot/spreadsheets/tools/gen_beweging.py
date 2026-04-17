"""Lightfoot — beweging_sport.xlsx generator."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from helpers import *
from openpyxl import Workbook

OUT = "/home/user/ekistic/lightfoot/spreadsheets/beweging_sport.xlsx"
EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'

def filosofie(wb):
    ws = mk_ws(wb, "Filosofie", "1A2B4A")
    ws.column_dimensions["A"].width = 90
    title_block(ws, 1, "BEWEGING & SPORT — Filosofie & Ontwerpprincipes",
                "Lightfoot · Project Existenzminimum 2.0 · Vlaanderen 2026")
    nota_blok(ws, 3, [
        "# BEWEGING & SPORT",
        "",
        "## Ontwerpdoel",
        "Dagelijkse beweging en krachttraining integreren in het minimummodel zonder",
        "lidmaatschapskosten, reistijd of extra ruimtebeslag.",
        "",
        "## Standaard bewegingssysteem",
        "→ Dagelijkse wandeling: 30-45 min buiten — primaire bijdrage aan gezondheid en mentale rust.",
        "→ Thuistraining (3× per week, 30-45 min): kettlebell + pull-up balk + lichaamsgewicht.",
        "→ Fietsen: dagelijkse verplaatsing — vervangt aparte cardiosessies.",
        "",
        "## Standaard trainingsprotocol",
        "→ Maandag: kettlebell swings (5×20) + goblet squats (4×12) + push-ups (4×15).",
        "→ Woensdag: pull-ups (5×max) + kettlebell clean & press (4×8) + hollow body holds.",
        "→ Vrijdag: Turkish get-up (3×3 per kant) + kettlebell deadlift (4×12) + dips.",
        "→ Zaterdag/zondag: wandeling 60 min of fietstocht.",
        "",
        "## Materiaal",
        "→ Kettlebell 24 kg (norm voor man, beginners 16 kg) — 1 gewicht is voldoende.",
        "→ Pull-up balk (deurframe) — geen schroeven, geen schade.",
        "→ Yogamat 6mm — vloertraining en stretching.",
        "→ Sportkleding: 2 sets functionele kleding (deelt was met kledingfunctie).",
        "",
        "## Uitzondering: Basic-Fit lidmaatschap",
        "→ Prijs: €22.99/maand (single club) — uitzondering bij gebrek aan thuisruimte.",
        "→ Voordeel: zwembad, groepslessen, extra toestellen.",
        "→ Nadeel: tijdsverlies reistijd, sociale drempel, maandelijkse verplichting.",
        "→ Keuze: thuismodel standaard. Basic-Fit als expliciete uitzondering bij bewuste voorkeur.",
        "",
        "## Sportkledingcapsule",
        "→ 2× functioneel t-shirt (polyester/nylon quick-dry) — Decathlon.",
        "→ 2× trainingsshort of legging — Decathlon.",
        "→ 1× sportsokken set (3 paar) — Decathlon.",
        "→ Sportschoenen: geen extra — New Balance 990 (dagelijks) voldoet voor thuistraining.",
        "→ Buitenloop: indien van toepassing — apart budgetteren als uitzondering.",
    ])

def inventaris(wb):
    ws = mk_ws(wb, "Inventaris", "2D4A7A")
    set_cols(ws, [28, 34, 22, 6, 10, 10, 12, 12])
    title_block(ws, 1, "Inventaris — Beweging & Sport")
    hdr(ws, 4, ["Item", "Specificatie", "Merk / Model (standaard)", "Qty",
                "Prijs/stuk (€)", "Totaal (€)", "Levensduur", "Afschr/mnd (€)"])

    items = [
        ("Kettlebell", "24 kg, gegoten ijzer, vinyl coating", "Decathlon 24 kg", 1, 55, 240),
        ("Pull-up balk (deurframe)", "Telescopisch, max. 100 kg", "Iron Gym Total Upper Body", 1, 30, 60),
        ("Yogamat", "6mm TPE, 183×61cm, anti-slip", "Decathlon Kimjaly 6mm", 1, 20, 48),
        ("Trainingshandschoenen", "Open vinger, anti-blaren", "Decathlon", 1, 12, 36),
        ("Functioneel t-shirt sport", "Quick-dry polyester", "Decathlon Kalenji", 2, 8, 24),
        ("Trainingsshort", "Stretch, 2-in-1 met voering", "Decathlon", 2, 12, 36),
        ("Sportsokken set", "Low-cut, 3 paar", "Decathlon", 1, 8, 18),
        ("Foam roller", "Hard EVA, 33×15cm", "Decathlon", 1, 15, 120),
    ]

    r = 5
    totaal_inv = 0
    totaal_afschr = 0
    for item, spec, merk, qty, prijs, levmnd in items:
        totaal = qty * prijs
        afschr = round(totaal / levmnd, 2)
        totaal_inv += totaal
        totaal_afschr += afschr
        drow(ws, r, [item, spec, merk, qty, prijs, totaal,
                     f"{levmnd // 12} jaar", afschr],
             fmts=[None, None, None, None, EUR, EUR0, None, EUR])
        r += 1

    trow(ws, r, ["TOTAAL", "", "", "", "", totaal_inv, "", round(totaal_afschr, 2)])
    ws.cell(r, 6).number_format = EUR0
    ws.cell(r, 8).number_format = EUR

def trainingsritme(wb):
    ws = mk_ws(wb, "Trainingsritme", "2D4A7A")
    set_cols(ws, [14, 20, 40, 14, 20])
    title_block(ws, 1, "Trainingsritme — Wekelijks schema")
    hdr(ws, 3, ["Dag", "Type", "Oefeningen / Activiteit", "Duur (min)", "Locatie"])

    schema = [
        ("Maandag", "Kracht (kettlebell)", "Swings 5×20 / Goblet squat 4×12 / Push-ups 4×15", 40, "Thuis"),
        ("Dinsdag", "Actief herstel", "Wandeling 30-45 min", 40, "Buiten"),
        ("Woensdag", "Kracht (pull/push)", "Pull-ups 5×max / KB clean&press 4×8 / Hollow body", 40, "Thuis"),
        ("Donderdag", "Rust of wandeling", "Optioneel: wandeling 30 min", 30, "Buiten / Rust"),
        ("Vrijdag", "Kracht (hinge/core)", "Turkish get-up 3×3 / KB deadlift 4×12 / Dips", 40, "Thuis"),
        ("Zaterdag", "Duurzame activiteit", "Fietstocht of wandeling 60+ min", 60, "Buiten"),
        ("Zondag", "Rust + mobiliteit", "Stretching + foam rolling", 20, "Thuis"),
    ]

    for r, (dag, typ, oefen, duur, loc) in enumerate(schema, 4):
        drow(ws, r, [dag, typ, oefen, duur, loc])

    trow(ws, 11, ["TOTAAL / WEEK", "", "", sum(s[3] for s in schema), ""])

def budget(wb):
    ws = mk_ws(wb, "Budget", "1A2B4A")
    set_cols(ws, [36, 18, 18, 28])
    title_block(ws, 1, "Budget — Beweging & Sport (TCO-overzicht)")

    ws.cell(3, 1, "INVESTERING").font = fB(col=C["navy"])
    hdr(ws, 4, ["Post", "Bedrag (€)", "", "Toelichting"])
    drow(ws, 5, ["Thuistraining materiaal (8 items)", 182, "", "Eenmalige aankoopkost"])
    trow(ws, 6, ["Totaal investering", 182, "", ""])
    ws.cell(6, 2).number_format = EUR0

    ws.cell(8, 1, "MAANDELIJKSE KOST — STANDAARD (thuis)").font = fB(col=C["navy"])
    hdr(ws, 9, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"])
    maand_items = [
        ("Afschrijving materiaal (thuis)", 3.83, "Op basis van levensduur per item"),
        ("Sportkleding was (gedeeld met kleding)", 0.50, "Extra wasbeurt per maand"),
    ]
    r = 10
    for post, m, toel in maand_items:
        drow(ws, r, [post, m, round(m * 12, 2), toel],
             fmts=[None, EUR, EUR0, None])
        r += 1

    totm = sum(x[1] for x in maand_items)
    trow(ws, r, ["TOTAAL / MAAND (thuis)", round(totm, 2), round(totm * 12, 2), ""])
    ws.cell(r, 2).number_format = EUR
    ws.cell(r, 3).number_format = EUR0

    ws.cell(r+2, 1, "UITZONDERING: Basic-Fit lidmaatschap").font = fB(col=C["navy_mid"])
    hdr(ws, r+3, ["Kostenpost", "Maandkost (€)", "Jaarkost (€)", "Toelichting"],
        bg=C["amber"])
    drow(ws, r+4, ["Basic-Fit (single club)", 22.99, round(22.99 * 12, 2),
                   "Vervangt thuismodel — uitzondering"],
         bg=C["amber"],
         fmts=[None, EUR, EUR0, None])

    ws.cell(r+6, 1, "TIJDSBELASTING").font = fB(col=C["navy"])
    hdr(ws, r+7, ["Activiteit", "Tijd/maand (min)", "Tijd/jaar (u)", ""])
    tbelast = [
        ("Krachtsessies (3× 40 min)", 120, ""),
        ("Wandelingen dagelijks (30 min avg)", 210, ""),
        ("Stretching + herstel", 40, ""),
    ]
    r2 = r + 8
    for act, min_mnd, _ in tbelast:
        drow(ws, r2, [act, min_mnd, round(min_mnd * 12 / 60, 1), ""])
        r2 += 1
    totmin = sum(x[1] for x in tbelast)
    trow(ws, r2, ["TOTAAL", totmin, round(totmin * 12 / 60, 1), ""])

def afhankelijkheden(wb):
    ws = mk_ws(wb, "Afhankelijkheden", "2D4A7A")
    set_cols(ws, [24, 20, 56])
    title_block(ws, 1, "Afhankelijkheden — Beweging & Sport")
    hdr(ws, 3, ["Van functie", "Naar functie", "Beschrijving"])
    afh = [
        ("Beweging", "Wonen / Shelter", "Pull-up balk: deurframe min. 70 cm breed, draagkracht 100 kg."),
        ("Beweging", "Wonen / Shelter", "Ruimte 2×2m voor kettlebell + mat vereist."),
        ("Beweging", "Kleding", "Sportkleding deelt wasroutine — geen extra wasmachine-budget."),
        ("Beweging", "Mobiliteit", "Fietsen telt als bewegingsactiviteit — geen dubbeltelling nodig."),
        ("Beweging", "Werk", "Bewegingspauzes (2× 5 min/uur) preventief bij zittend werk."),
    ]
    for r, (van, naar, beschr) in enumerate(afh, 4):
        drow(ws, r, [van, naar, beschr])

def main():
    wb = Workbook()
    wb.remove(wb.active)
    filosofie(wb)
    inventaris(wb)
    trainingsritme(wb)
    budget(wb)
    afhankelijkheden(wb)
    wb.save(OUT)
    print(f"Opgeslagen: {OUT}")

if __name__ == "__main__":
    main()
